"""
batch5_master.py — Master Summary Script
Reads all batch1-4 outputs, builds master school profile, top20 lists,
state summary, color-coded MASTER_SUMMARY.xlsx, and dashboard_data.json.
"""

import os
import json
import warnings
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime

warnings.filterwarnings("ignore")

# ── Paths ──────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).resolve().parent.parent
OUT_DIR    = BASE_DIR / "outputs"
L1         = OUT_DIR / "L1_STATE"
L2         = OUT_DIR / "L2_DISTRICT"
L3         = OUT_DIR / "L3_BLOCK"
L4         = OUT_DIR / "L4_SCHOOL"
L5         = OUT_DIR / "L5_MASTER"

L5.mkdir(parents=True, exist_ok=True)


# ── Helpers ────────────────────────────────────────────────────────────────
def _load(path, **kwargs):
    """Load CSV, return empty DataFrame if file missing."""
    p = Path(path)
    if not p.exists():
        print(f"  [WARN] Missing: {p.name}")
        return pd.DataFrame()
    df = pd.read_csv(p, dtype=str, **kwargs)
    return df


def _norm_id(df, col="school_id"):
    """Normalise school_id: strip, lower, keep as str (leading zeros preserved)."""
    if col in df.columns:
        df[col] = df[col].astype(str).str.strip().str.lower()
    return df


def _to_title_id(df):
    """Batch2-4 use 'School_ID' — rename to school_id and normalise."""
    if "School_ID" in df.columns:
        df = df.rename(columns={
            "School_ID": "school_id",
            "School_Name": "school_name",
            "District": "district",
            "Block": "block",
        })
    return _norm_id(df)


def _num(df, cols):
    """Convert listed cols to numeric, coerce errors to NaN."""
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def _safe_pct(val):
    if pd.isna(val):
        return None
    return round(float(val), 2)


# ══════════════════════════════════════════════════════════════════════════
# STEP 1 — Load all batch outputs
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("BATCH 5 — MASTER SUMMARY")
print("="*60)
print("\n[STEP 1] Loading batch outputs …")

# ── Batch 1 (school-level, lowercase ids) ─────────────────────────────────
b1_growth   = _norm_id(_load(L4 / "01_growth_decline.csv"),   "school_id")
b1_stab     = _norm_id(_load(L4 / "02_stability_index.csv"),  "school_id")
b1_gender   = _norm_id(_load(L4 / "03_gender_equity.csv"),    "school_id")
b1_cont     = _norm_id(_load(L4 / "04_continuity_index.csv"), "school_id")
b1_dist     = _norm_id(_load(L4 / "05_class_distribution.csv"), "school_id")

# ── Batch 2 (school/block/district, Title case ids) ───────────────────────
b2_risk     = _to_title_id(_load(L4 / "08_school_risk_index.csv"))
b2_block    = _load(L3 / "06_block_health.csv")
b2_district = _load(L2 / "07_district_performance.csv")

# ── Batch 3 (school-level, Title case ids) ────────────────────────────────
b3_driver   = _to_title_id(_load(L4 / "11_growth_driver.csv"))
b3_decline  = _to_title_id(_load(L4 / "12_decline_reason.csv"))
b3_potential= _to_title_id(_load(L4 / "13_enrollment_potential.csv"))
b3_warning  = _to_title_id(_load(L4 / "14_early_warning.csv"))
b3_segment  = _to_title_id(_load(L4 / "15_segment_split.csv"))

# ── Batch 4 (school-level, Title case ids) ────────────────────────────────
b4_girls    = _to_title_id(_load(L4 / "21_girls_ladder.csv"))
b4_recovery = _to_title_id(_load(L4 / "18_recovery_index.csv"))
b4_small    = _to_title_id(_load(L4 / "19_small_school.csv"))

# ── L1 state ──────────────────────────────────────────────────────────────
b1_stype    = _load(L1 / "01b_schooltype_trend.csv")

# ── Pre-existing L5 master files ──────────────────────────────────────────
l5_highrisk  = _to_title_id(_load(L5 / "high_risk_schools.csv"))
l5_model     = _to_title_id(_load(L5 / "model_schools.csv"))
l5_invest    = _to_title_id(_load(L5 / "investment_worthy_schools.csv"))
l5_recovery  = _to_title_id(_load(L5 / "recovery_schools.csv"))
l5_closure   = _to_title_id(_load(L5 / "closure_risk_schools.csv"))
l5_alerts    = _load(L5 / "early_warning_summary.csv")
l5_critical  = _to_title_id(_load(L5 / "critical_alerts.csv"))

print("  Loaded all CSVs.")


# ══════════════════════════════════════════════════════════════════════════
# STEP 2 — Build master school profile table
# ══════════════════════════════════════════════════════════════════════════
print("\n[STEP 2] Building master school profile …")

# Base: batch1 growth (has every school)
master = b1_growth[["school_id", "school_name", "district", "block",
                     "School_Type", "Class_Range_Label",
                     "Enroll_2223", "Enroll_2324", "Enroll_2425", "Enroll_2526",
                     "Overall_Growth", "Trend_Tag"]].copy()
master = master.rename(columns={"Overall_Growth": "Overall_Growth_pct"})

# Stability
if not b1_stab.empty:
    master = master.merge(
        b1_stab[["school_id", "Stability_Label", "CV_Pct"]],
        on="school_id", how="left"
    )

# Gender equity
if not b1_gender.empty:
    master = master.merge(
        b1_gender[["school_id", "Girls_Ratio_Current", "Girls_Growth_Pct",
                   "Boys_Growth_Pct", "Girls_Imbalance_Flag", "Girls_Declining_Flag"]],
        on="school_id", how="left"
    )

# Continuity
if not b1_cont.empty:
    master = master.merge(
        b1_cont[["school_id", "Avg_Continuity_Current",
                 "Low_Continuity_Flag", "Weak_Continuity_Flag"]],
        on="school_id", how="left"
    )

# Class distribution
if not b1_dist.empty:
    master = master.merge(
        b1_dist[["school_id", "Enroll_Gap_Current", "Enroll_Pattern_Tag"]],
        on="school_id", how="left"
    )

# Risk index (batch2) — normalise ids already done by _to_title_id
if not b2_risk.empty:
    master = master.merge(
        b2_risk[["school_id", "Risk_Score", "Risk_Level",
                 "Decline_flag", "Severe_Decline_flag",
                 "Low_Continuity_flag", "Wide_Enroll_Spread_flag",
                 "Low_Enrollment_flag", "Volatile_flag",
                 "Girls_Decline_flag", "Gender_Imbalance_flag",
                 "Small_School_flag"]],
        on="school_id", how="left"
    )

# Growth driver (batch3) — only high-growth schools; left join
if not b3_driver.empty:
    master = master.merge(
        b3_driver[["school_id", "Gender_Driver", "Class_Driver",
                   "Growth_Pattern_Tag", "Model_School_Flag"]],
        on="school_id", how="left"
    )
    master["Model_School_Flag"] = master["Model_School_Flag"].fillna("False")

# Decline reason (batch3) — only declining schools; left join
if not b3_decline.empty:
    master = master.merge(
        b3_decline[["school_id", "Primary_Reason", "Confidence_Level",
                    "All_Reasons_Detected"]],
        on="school_id", how="left"
    )

# Enrollment potential (batch3)
if not b3_potential.empty:
    master = master.merge(
        b3_potential[["school_id", "Potential_Score", "Potential_Tag",
                      "Investment_Worthy"]],
        on="school_id", how="left"
    )

# Girls enrollment progression (batch4)
if not b4_girls.empty:
    master = master.merge(
        b4_girls[["school_id", "Girls_Progression_pct",
                  "Cont_C8_C9_pct", "Weakest_Step",
                  "C10_C11_Trend", "Ladder_Tag"]],
        on="school_id", how="left"
    )

# Recovery (batch4)
if not b4_recovery.empty:
    master = master.merge(
        b4_recovery[["school_id", "Recovery_Pattern", "Recovery_Strength",
                     "Had_Decline"]],
        on="school_id", how="left"
    )

# Small school / closure risk (batch4)
if not b4_small.empty:
    master = master.merge(
        b4_small[["school_id", "Size_Tag", "Size_Trend",
                  "Closure_Risk_Score", "Closure_Risk_Tag"]],
        on="school_id", how="left"
    )

# Early warning flags per school — pivot to get unique alerts per school
if not b3_warning.empty:
    warn_agg = (b3_warning.groupby("school_id")["Alert_Type"]
                .apply(lambda x: "; ".join(sorted(x.unique())))
                .reset_index()
                .rename(columns={"Alert_Type": "Active_Alerts"}))
    warn_severity = (b3_warning.groupby("school_id")["Severity"]
                     .apply(lambda x: "CRITICAL" if "CRITICAL" in x.values
                            else ("HIGH" if "HIGH" in x.values else "LOW"))
                     .reset_index()
                     .rename(columns={"Severity": "Max_Alert_Severity"}))
    warn_count = (b3_warning.groupby("school_id").size()
                  .reset_index(name="Alert_Count"))
    warn_df = warn_agg.merge(warn_severity, on="school_id").merge(warn_count, on="school_id")
    master = master.merge(warn_df, on="school_id", how="left")
    master["Alert_Count"] = master["Alert_Count"].fillna(0).astype(int)

# Numeric conversions
num_cols = [
    "Enroll_2223", "Enroll_2324", "Enroll_2425", "Enroll_2526",
    "Overall_Growth_pct", "CV_Pct", "Girls_Ratio_Current",
    "Girls_Growth_Pct", "Boys_Growth_Pct", "Avg_Continuity_Current",
    "Enroll_Gap_Current", "Risk_Score", "Potential_Score",
    "Girls_Progression_pct", "Cont_C8_C9_pct",
    "Closure_Risk_Score", "Alert_Count",
]
master = _num(master, num_cols)

print(f"  Master table: {len(master):,} schools, {master.columns.size} columns")


# ══════════════════════════════════════════════════════════════════════════
# STEP 3 — Top 20 lists
# ══════════════════════════════════════════════════════════════════════════
print("\n[STEP 3] Building Top 20 lists …")

id_cols = ["school_id", "school_name", "district", "block", "School_Type"]

# Top 20 growth schools
top20_growth = (master[master["Overall_Growth_pct"].notna()]
                .sort_values("Overall_Growth_pct", ascending=False)
                .head(20)[id_cols + ["Enroll_2223", "Enroll_2526",
                                     "Overall_Growth_pct", "Trend_Tag"]]
                .reset_index(drop=True))
top20_growth.index += 1

# Top 20 decline schools
top20_decline = (master[master["Overall_Growth_pct"].notna()]
                 .sort_values("Overall_Growth_pct", ascending=True)
                 .head(20)[id_cols + ["Enroll_2223", "Enroll_2526",
                                      "Overall_Growth_pct", "Primary_Reason"]]
                 .reset_index(drop=True))
top20_decline.index += 1

# Top 20 high-risk schools
top20_risk = (master[master["Risk_Level"] == "HIGH RISK"]
              .sort_values("Risk_Score", ascending=False)
              .head(20)[id_cols + ["Risk_Score", "Risk_Level",
                                   "Overall_Growth_pct", "Active_Alerts"]]
              .reset_index(drop=True))
top20_risk.index += 1

# Top 20 model schools (high growth + Model_School_Flag)
top20_model = (master[master["Model_School_Flag"].astype(str).str.lower() == "true"]
               .sort_values("Overall_Growth_pct", ascending=False)
               .head(20)[id_cols + ["Overall_Growth_pct", "Growth_Pattern_Tag",
                                    "Girls_Growth_Pct"]]
               .reset_index(drop=True))
top20_model.index += 1

# Top 20 investment-worthy schools
if "Investment_Worthy" in master.columns:
    top20_invest = (master[master["Investment_Worthy"].astype(str).str.lower() == "true"]
                    .sort_values("Potential_Score", ascending=False)
                    .head(20)[id_cols + ["Enroll_2526", "Potential_Score",
                                         "Potential_Tag", "Overall_Growth_pct"]]
                    .reset_index(drop=True))
    top20_invest.index += 1
else:
    top20_invest = pd.DataFrame()

# Red-zone blocks (Block_Zone == "RED")
if not b2_block.empty:
    red_blocks = (b2_block[b2_block["Block_Zone"] == "RED"]
                  .sort_values("Avg_Growth_pct", ascending=True)
                  [["District", "Block", "Total_Schools", "Decline_Schools",
                    "Low_Continuity_Schools", "Girls_Declining_Schools",
                    "Avg_Growth_pct", "Block_Zone", "Priority_Action_Flag"]]
                  .reset_index(drop=True))
else:
    red_blocks = pd.DataFrame()

print(f"  Growth top20: {len(top20_growth)}, Decline top20: {len(top20_decline)}, "
      f"High-risk top20: {len(top20_risk)}, Model top20: {len(top20_model)}")
print(f"  Red-zone blocks: {len(red_blocks)}")


# ══════════════════════════════════════════════════════════════════════════
# STEP 4 — State summary
# ══════════════════════════════════════════════════════════════════════════
print("\n[STEP 4] Building state summary …")

enroll_cols = ["Enroll_2223", "Enroll_2324", "Enroll_2425", "Enroll_2526"]
master = _num(master, enroll_cols)

total_schools  = len(master)
total_enroll   = {c: int(master[c].sum()) for c in enroll_cols}
overall_growth = _safe_pct((total_enroll["Enroll_2526"] - total_enroll["Enroll_2223"])
                           / total_enroll["Enroll_2223"] * 100)

high_risk_count   = int((master["Risk_Level"] == "HIGH RISK").sum())  if "Risk_Level" in master.columns else 0
medium_risk_count = int((master["Risk_Level"] == "MEDIUM RISK").sum()) if "Risk_Level" in master.columns else 0
low_risk_count    = int((master["Risk_Level"] == "LOW RISK").sum())   if "Risk_Level" in master.columns else 0

growing_count  = int((master["Overall_Growth_pct"] > 0).sum())
declining_count= int((master["Overall_Growth_pct"] < 0).sum())

model_count    = int((master["Model_School_Flag"].astype(str).str.lower() == "true").sum()) if "Model_School_Flag" in master.columns else 0
invest_count   = int((master["Investment_Worthy"].astype(str).str.lower() == "true").sum()) if "Investment_Worthy" in master.columns else 0
closure_count  = int((master["Closure_Risk_Tag"].isin(["High Risk", "CRITICAL"])).sum()) if "Closure_Risk_Tag" in master.columns else 0

districts_count = master["district"].nunique()
blocks_count    = master["block"].nunique()

avg_girls_ratio = _safe_pct(master["Girls_Ratio_Current"].mean() * 100) if "Girls_Ratio_Current" in master.columns else None

# Critical alerts summary
critical_alert_count = 0
if not l5_alerts.empty and "Severity" in l5_alerts.columns:
    critical_alert_count = int((l5_alerts["Severity"] == "CRITICAL").sum())

state_kpis = {
    "Report_Date":          datetime.today().strftime("%Y-%m-%d"),
    "Total_Schools":        total_schools,
    "Total_Districts":      districts_count,
    "Total_Blocks":         blocks_count,
    "Total_Enrollment_2223": total_enroll["Enroll_2223"],
    "Total_Enrollment_2324": total_enroll["Enroll_2324"],
    "Total_Enrollment_2425": total_enroll["Enroll_2425"],
    "Total_Enrollment_2526": total_enroll["Enroll_2526"],
    "Overall_Enrollment_Growth_pct": overall_growth,
    "Growing_Schools":      growing_count,
    "Declining_Schools":    declining_count,
    "High_Risk_Schools":    high_risk_count,
    "Medium_Risk_Schools":  medium_risk_count,
    "Low_Risk_Schools":     low_risk_count,
    "Model_Schools":        model_count,
    "Investment_Worthy_Schools": invest_count,
    "Closure_Risk_Schools": closure_count,
    "Avg_Girls_Ratio_pct":  avg_girls_ratio,
    "Critical_Alert_Types": critical_alert_count,
    "Red_Zone_Blocks":      len(red_blocks),
}

state_kpis_df = pd.DataFrame([state_kpis]).T.reset_index()
state_kpis_df.columns = ["KPI", "Value"]
state_kpis_df.to_csv(L1 / "state_kpis.csv", index=False)

# State enrollment trend (by year)
enroll_trend = pd.DataFrame({
    "Year":             ["2022-23", "2023-24", "2024-25", "2025-26"],
    "Total_Enrollment": [total_enroll[c] for c in enroll_cols],
    "Growing_Schools":  [
        int((master["Growth_2223_2324"] > 0).sum()) if "Growth_2223_2324" in master.columns else 0,
        int((master["Growth_2324_2425"] > 0).sum()) if "Growth_2324_2425" in master.columns else 0,
        int((master["Growth_2425_2526"] > 0).sum()) if "Growth_2425_2526" in master.columns else 0,
        growing_count,
    ],
})
if "Girls_2223" in master.columns:
    enroll_trend["Girls_Enrollment"] = [
        int(master["Girls_2223"].sum()),
        int(master["Girls_2324"].sum()) if "Girls_2324" in master.columns else 0,
        int(master["Girls_2425"].sum()) if "Girls_2425" in master.columns else 0,
        int(master["Girls_2526"].sum()) if "Girls_2526" in master.columns else 0,
    ]
enroll_trend.to_csv(L1 / "state_enrollment_trend.csv", index=False)

# District summary (from batch2)
if not b2_district.empty:
    b2_district.to_csv(L5 / "district_summary.csv", index=False)

print(f"  State KPIs saved. Total enrollment 2022-23: {total_enroll['Enroll_2223']:,} "
      f"-> 2025-26: {total_enroll['Enroll_2526']:,} ({overall_growth:+.1f}%)")


# ══════════════════════════════════════════════════════════════════════════
# STEP 5 — MASTER_SUMMARY.xlsx (color-coded, 9 sheets)
# ══════════════════════════════════════════════════════════════════════════
print("\n[STEP 5] Building MASTER_SUMMARY.xlsx …")

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    # ── Color palette ──────────────────────────────────────────────────────
    CLR = {
        "header_dark":  "1F3864",   # dark navy  — header background
        "header_light": "FFFFFF",   # white text on header
        "green_dark":   "1E8449",
        "green_mid":    "A9DFBF",
        "green_light":  "EAFAF1",
        "red_dark":     "922B21",
        "red_mid":      "F1948A",
        "red_light":    "FDEDEC",
        "amber_dark":   "9C640C",
        "amber_mid":    "FAD7A0",
        "amber_light":  "FEF9E7",
        "blue_light":   "EBF5FB",
        "purple_light": "F5EEF8",
        "grey_light":   "F2F3F4",
        "white":        "FFFFFF",
    }

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color="000000", size=10):
        return Font(bold=bold, color=color, size=size)

    def _border():
        thin = Side(style="thin", color="D0D3D4")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _write_sheet(ws, df, header_color="1F3864", freeze=True,
                     col_widths=None, row_colors=None):
        """Write a DataFrame to a worksheet with formatting."""
        # Header row
        ws.append(list(df.columns))
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.fill    = _fill(header_color)
            cell.font    = _font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border  = _border()

        # Data rows
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = _border()
                cell.alignment = Alignment(vertical="center")
                # Default alternating row color
                base_fill = CLR["white"] if r_idx % 2 == 0 else CLR["grey_light"]
                cell.fill = _fill(base_fill)

        # Custom row coloring
        if row_colors:
            for r_idx in range(2, ws.max_row + 1):
                row_color = row_colors(ws, r_idx, df)
                if row_color:
                    for c_idx in range(1, len(df.columns) + 1):
                        ws.cell(row=r_idx, column=c_idx).fill = _fill(row_color)

        # Column widths
        for c_idx, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(c_idx)
            if col_widths and col in col_widths:
                ws.column_dimensions[col_letter].width = col_widths[col]
            else:
                max_len = max(len(str(col)), 10)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        if freeze:
            ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Sheet 1: State KPIs ────────────────────────────────────────────────
    ws1 = wb.create_sheet("1_State_KPIs")
    ws1.sheet_view.showGridLines = False
    _write_sheet(ws1, state_kpis_df, header_color=CLR["header_dark"],
                 col_widths={"KPI": 35, "Value": 25})
    ws1["A1"].value = "KPI"
    ws1["B1"].value = "Value"

    # ── Sheet 2: School Type Trend ─────────────────────────────────────────
    ws2 = wb.create_sheet("2_SchoolType_Trend")
    if not b1_stype.empty:
        def _stype_color(ws, r_idx, df):
            tag_col = list(df.columns).index("Type_Health_Tag") if "Type_Health_Tag" in df.columns else -1
            if tag_col >= 0:
                val = ws.cell(row=r_idx, column=tag_col+1).value
                if val == "Thriving":   return CLR["green_mid"]
                if val == "Growing":    return CLR["green_light"]
                if val == "Critical":   return CLR["red_mid"]
                if val == "Declining":  return CLR["red_light"]
                if val == "Mixed":      return CLR["amber_light"]
            return None
        _write_sheet(ws2, b1_stype, header_color="2E4057",
                     row_colors=_stype_color)

    # ── Sheet 3: District Performance ─────────────────────────────────────
    ws3 = wb.create_sheet("3_District_Performance")
    if not b2_district.empty:
        b2_dist_disp = b2_district.copy()
        def _dist_color(ws, r_idx, df):
            tag_col = list(df.columns).index("District_Health_Tag") if "District_Health_Tag" in df.columns else -1
            if tag_col >= 0:
                val = ws.cell(row=r_idx, column=tag_col+1).value
                if val == "Strong":   return CLR["green_light"]
                if val == "Moderate": return CLR["amber_light"]
                if val == "Weak":     return CLR["red_light"]
            return None
        _write_sheet(ws3, b2_dist_disp, header_color="1B4F72",
                     row_colors=_dist_color)

    # ── Sheet 4: Top 20 Growth Schools ────────────────────────────────────
    ws4 = wb.create_sheet("4_Top20_Growth")
    if not top20_growth.empty:
        top20_growth_disp = top20_growth.reset_index().rename(columns={"index": "Rank"})
        def _growth_color(ws, r_idx, df):
            g_col = list(df.columns).index("Overall_Growth_pct") if "Overall_Growth_pct" in df.columns else -1
            if g_col >= 0:
                try:
                    val = float(ws.cell(row=r_idx, column=g_col+1).value or 0)
                    if val >= 100: return CLR["green_dark"]
                    if val >= 50:  return CLR["green_mid"]
                    if val >= 20:  return CLR["green_light"]
                except Exception:
                    pass
            return None
        _write_sheet(ws4, top20_growth_disp, header_color=CLR["green_dark"],
                     row_colors=_growth_color)

    # ── Sheet 5: Top 20 Decline Schools ───────────────────────────────────
    ws5 = wb.create_sheet("5_Top20_Decline")
    if not top20_decline.empty:
        top20_decline_disp = top20_decline.reset_index().rename(columns={"index": "Rank"})
        def _decline_color(ws, r_idx, df):
            g_col = list(df.columns).index("Overall_Growth_pct") if "Overall_Growth_pct" in df.columns else -1
            if g_col >= 0:
                try:
                    val = float(ws.cell(row=r_idx, column=g_col+1).value or 0)
                    if val <= -50: return CLR["red_mid"]
                    if val <= -25: return CLR["red_light"]
                    if val <= -10: return CLR["amber_light"]
                except Exception:
                    pass
            return None
        _write_sheet(ws5, top20_decline_disp, header_color=CLR["red_dark"],
                     row_colors=_decline_color)

    # ── Sheet 6: High Risk Schools ────────────────────────────────────────
    ws6 = wb.create_sheet("6_High_Risk_Schools")
    if not top20_risk.empty:
        top20_risk_disp = top20_risk.reset_index().rename(columns={"index": "Rank"})
        def _risk_color(ws, r_idx, df):
            r_col = list(df.columns).index("Risk_Score") if "Risk_Score" in df.columns else -1
            if r_col >= 0:
                try:
                    val = float(ws.cell(row=r_idx, column=r_col+1).value or 0)
                    if val >= 10: return CLR["red_mid"]
                    if val >= 7:  return CLR["amber_mid"]
                    if val >= 4:  return CLR["amber_light"]
                except Exception:
                    pass
            return None
        _write_sheet(ws6, top20_risk_disp, header_color=CLR["red_dark"],
                     row_colors=_risk_color)

    # ── Sheet 7: Red Zone Blocks ───────────────────────────────────────────
    ws7 = wb.create_sheet("7_Red_Zone_Blocks")
    if not red_blocks.empty:
        _write_sheet(ws7, red_blocks.reset_index(drop=True), header_color="922B21",
                     row_colors=lambda ws, r, df: CLR["red_light"])
    else:
        ws7.append(["No RED zone blocks found"])

    # ── Sheet 8: Model Schools ────────────────────────────────────────────
    ws8 = wb.create_sheet("8_Model_Schools")
    if not top20_model.empty:
        top20_model_disp = top20_model.reset_index().rename(columns={"index": "Rank"})
        _write_sheet(ws8, top20_model_disp, header_color=CLR["green_dark"],
                     row_colors=lambda ws, r, df: CLR["green_light"])

    # ── Sheet 9: Master School Profile ────────────────────────────────────
    ws9 = wb.create_sheet("9_Master_Profile")
    master_disp = master.copy()

    def _master_color(ws, r_idx, df):
        if "Risk_Level" in df.columns:
            rl_col = list(df.columns).index("Risk_Level")
            val = ws.cell(row=r_idx, column=rl_col+1).value
            if val == "HIGH RISK":   return CLR["red_light"]
            if val == "MEDIUM RISK": return CLR["amber_light"]
        return None

    _write_sheet(ws9, master_disp, header_color=CLR["header_dark"],
                 row_colors=_master_color)

    # Save
    xlsx_path = L5 / "MASTER_SUMMARY.xlsx"
    wb.save(xlsx_path)
    print(f"  Saved: {xlsx_path}")

except ImportError:
    print("  [WARN] openpyxl not available — skipping MASTER_SUMMARY.xlsx")

# ── Individual top20 CSVs ─────────────────────────────────────────────────
if not top20_growth.empty:
    top20_growth.to_csv(L5 / "top20_growth_schools.csv", index=False)
if not top20_decline.empty:
    top20_decline.to_csv(L5 / "top20_decline_schools.csv", index=False)
if not top20_risk.empty:
    top20_risk.to_csv(L5 / "top20_high_risk_schools.csv", index=False)
if not red_blocks.empty:
    red_blocks.to_csv(L5 / "red_zone_blocks.csv", index=False)

# ── Same-condition clusters ───────────────────────────────────────────────
# Group schools sharing the same Risk_Level + Trend_Tag + Enroll_Pattern_Tag
# (clusters of 3+ schools indicate a systemic / geographic pattern)
cluster_keys = [k for k in ["Risk_Level", "Trend_Tag", "Enroll_Pattern_Tag", "School_Type"]
                if k in master.columns]
if cluster_keys:
    grp = (master.dropna(subset=cluster_keys)
           .groupby(cluster_keys, as_index=False)
           .agg(
               School_Count=("school_id", "count"),
               Districts=("district", lambda x: ", ".join(sorted(x.dropna().unique())[:5])),
               Blocks=("block",   lambda x: ", ".join(sorted(x.dropna().unique())[:5])),
               Avg_Enrollment_2526=("Enroll_2526",      "mean"),
               Avg_Growth_pct=("Overall_Growth_pct", "mean"),
               Avg_Risk_Score=("Risk_Score",         "mean"),
           ))
    grp = grp[grp["School_Count"] >= 3].sort_values("School_Count", ascending=False).reset_index(drop=True)
    grp["Avg_Enrollment_2526"] = grp["Avg_Enrollment_2526"].round(1)
    grp["Avg_Growth_pct"]      = grp["Avg_Growth_pct"].round(2)
    grp["Avg_Risk_Score"]      = grp["Avg_Risk_Score"].round(1)
    grp.to_csv(L5 / "same_condition_clusters.csv", index=False)
    print(f"  Same-condition clusters: {len(grp)} groups (3+ schools each)")

# ── Master profile CSV ────────────────────────────────────────────────────
master.to_csv(L5 / "master_school_profiles.csv", index=False)
print(f"  Saved: {L5 / 'master_school_profiles.csv'}")


# ══════════════════════════════════════════════════════════════════════════
# STEP 6 — dashboard_data.json
# ══════════════════════════════════════════════════════════════════════════
print("\n[STEP 6] Building dashboard_data.json …")

def _df_to_records(df, max_rows=None):
    """Convert DataFrame to list of dicts with NaN cleaned."""
    if df.empty:
        return []
    if max_rows:
        df = df.head(max_rows)
    return [{k: (None if (isinstance(v, float) and np.isnan(v)) else v)
             for k, v in row.items()}
            for row in df.to_dict(orient="records")]


# ── Load heatmap CSVs (L3_BLOCK) ──────────────────────────────────────────
hmap_enroll = _load(L3 / "22_heatmap_enrollment.csv")
hmap_gender = _load(L3 / "22_heatmap_gender.csv")
hmap_trans  = _load(L3 / "22_heatmap_transition.csv")
hmap_risk   = _load(L3 / "22_heatmap_risk.csv")

# ── Load same_condition_clusters (already saved in L5) ───────────────────
clusters_df = _load(L5 / "same_condition_clusters.csv")

# ── Load additional L4 files for JSON ────────────────────────────────────
b4_recovery_full = _to_title_id(_load(L4 / "18_recovery_index.csv"))
b3_segment_full  = _to_title_id(_load(L4 / "15_segment_split.csv"))

# ── all_schools: master profile slim (key columns only) ──────────────────
all_schools_cols = [
    "school_id", "school_name", "district", "block",
    "School_Type", "Class_Range_Label",
    "Enroll_2223", "Enroll_2324", "Enroll_2425", "Enroll_2526",
    "Overall_Growth_pct", "Trend_Tag", "Stability_Label",
    "Girls_Ratio_Current", "Girls_Imbalance_Flag", "Girls_Declining_Flag",
    "Avg_Continuity_Current", "Low_Continuity_Flag",
    "Enroll_Pattern_Tag", "Risk_Score", "Risk_Level",
    "Model_School_Flag", "Investment_Worthy",
    "Closure_Risk_Tag", "Size_Tag",
    "Active_Alerts", "Alert_Count", "Max_Alert_Severity",
]
all_schools_df = master[[c for c in all_schools_cols if c in master.columns]].copy()

# ── zone_summary: count blocks per zone ───────────────────────────────────
zone_summary_json = {}
if not b2_block.empty and "Block_Zone" in b2_block.columns:
    zone_counts = b2_block["Block_Zone"].value_counts().to_dict()
    zone_summary_json = zone_counts

# ── gender_data: per-district girls ratio + trend ─────────────────────────
gender_data_json = []
if not b2_district.empty:
    gd_cols = [c for c in ["District", "Girls_Growth_pct", "Boys_Growth_pct",
                            "Girls_Declining_Schools", "Avg_Girls_Ratio_pct"]
               if c in b2_district.columns]
    if gd_cols:
        gd = b2_district[gd_cols].copy()
        gd = _num(gd, [c for c in gd_cols if c != "District"])
        gender_data_json = _df_to_records(gd)

# ── girls_ladder: per-school girls progression ────────────────────────────
girls_ladder_json = _df_to_records(b4_girls) if not b4_girls.empty else []

# ── transition_data: per-school continuity (slim) ─────────────────────────
transition_cols = ["school_id", "school_name", "district", "block",
                   "School_Type", "Avg_Continuity_Current",
                   "Continuity_8_9_2526", "Continuity_9_10_2526",
                   "Continuity_10_11_2526", "Continuity_11_12_2526",
                   "Low_Continuity_Flag", "Weak_Continuity_Flag"]
if not b1_cont.empty:
    tc = b1_cont[[c for c in transition_cols if c in b1_cont.columns]].copy()
    transition_json = _df_to_records(tc)
else:
    transition_json = []

# ── high_risk_schools: ALL high risk (not just top20) ─────────────────────
high_risk_json = []
if "Risk_Level" in master.columns:
    hr = master[master["Risk_Level"] == "HIGH RISK"].sort_values(
        "Risk_Score", ascending=False
    )[[c for c in ["school_id", "school_name", "district", "block",
                   "School_Type", "Enroll_2223", "Enroll_2526",
                   "Overall_Growth_pct", "Risk_Score", "Risk_Level",
                   "Active_Alerts"] if c in master.columns]]
    high_risk_json = _df_to_records(hr)

# ── early_warnings: per-school alert records ──────────────────────────────
early_warnings_json = _df_to_records(b3_warning) if not b3_warning.empty else []

# ── warning_summary: alert type summary ───────────────────────────────────
warning_summary_json = _df_to_records(l5_alerts) if not l5_alerts.empty else []

# ── recovery_schools ──────────────────────────────────────────────────────
recovery_json = []
if not b4_recovery_full.empty:
    rec = b4_recovery_full[b4_recovery_full["Had_Decline"].astype(str).str.lower() == "true"]
    recovery_json = _df_to_records(rec)

# ── model_schools ─────────────────────────────────────────────────────────
model_json = _df_to_records(l5_model) if not l5_model.empty else _df_to_records(top20_model)

# ── condition_clusters ────────────────────────────────────────────────────
clusters_json = _df_to_records(clusters_df) if not clusters_df.empty else []

# ── enrollment_trend (state level, 4 years) ──────────────────────────────
enroll_trend_json = [
    {"year": "2022-23", "total": total_enroll["Enroll_2223"]},
    {"year": "2023-24", "total": total_enroll["Enroll_2324"]},
    {"year": "2024-25", "total": total_enroll["Enroll_2425"]},
    {"year": "2025-26", "total": total_enroll["Enroll_2526"]},
]

# ── filter options ────────────────────────────────────────────────────────
stype_options = sorted(master["School_Type"].dropna().unique().tolist()) if "School_Type" in master.columns else []

# ── Assemble dashboard with ALL expected keys ─────────────────────────────
dashboard = {
    "metadata": {
        "generated_at":  datetime.today().strftime("%Y-%m-%d %H:%M"),
        "total_schools":  total_schools,
        "total_districts": districts_count,
        "total_blocks":   blocks_count,
        "data_years":     ["2022-23", "2023-24", "2024-25", "2025-26"],
    },
    "state_kpis":          state_kpis,
    "school_type_data":    _df_to_records(b1_stype) if not b1_stype.empty else [],
    "district_data":       _df_to_records(b2_district) if not b2_district.empty else [],
    "block_health":        _df_to_records(b2_block)    if not b2_block.empty    else [],
    "zone_summary":        zone_summary_json,
    "top20_growth":        _df_to_records(top20_growth),
    "top20_decline":       _df_to_records(top20_decline),
    "high_risk_schools":   high_risk_json,
    "early_warnings":      early_warnings_json,
    "warning_summary":     warning_summary_json,
    "gender_data":         gender_data_json,
    "girls_ladder":        girls_ladder_json,
    "transition_data":     transition_json,
    "segment_data":        _df_to_records(b3_segment_full) if not b3_segment_full.empty else [],
    "heatmap_enrollment":  _df_to_records(hmap_enroll) if not hmap_enroll.empty else [],
    "heatmap_gender":      _df_to_records(hmap_gender) if not hmap_gender.empty else [],
    "heatmap_transition":  _df_to_records(hmap_trans)  if not hmap_trans.empty  else [],
    "heatmap_risk":        _df_to_records(hmap_risk)   if not hmap_risk.empty   else [],
    "all_schools":         _df_to_records(all_schools_df),
    "recovery_schools":    recovery_json,
    "model_schools":       model_json,
    "condition_clusters":  clusters_json,
    # retained extras (for backward compat)
    "enrollment_trend":    enroll_trend_json,
    "filter_options": {
        "school_types": stype_options,
        "districts":    sorted(master["district"].dropna().unique().tolist()),
    },
}

json_path = OUT_DIR / "dashboard_data.json"
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(dashboard, f, ensure_ascii=False, indent=2, default=str)
print(f"  Saved: {json_path}")
print(f"  Keys written: {len(dashboard)}")


# ══════════════════════════════════════════════════════════════════════════
# STEP 7 — Final print summary
# ══════════════════════════════════════════════════════════════════════════
print("\n" + "="*60)
print("MASTER SUMMARY — UTTAR PRADESH GOVERNMENT SCHOOLS")
print(f"Report Date: {datetime.today().strftime('%d %B %Y')}")
print("="*60)

print(f"\n  COVERAGE")
print(f"    Schools : {total_schools:>8,}")
print(f"    Districts: {districts_count:>7,}")
print(f"    Blocks  : {blocks_count:>8,}")

print(f"\n  ENROLLMENT")
print(f"    2022-23 : {total_enroll['Enroll_2223']:>10,}")
print(f"    2023-24 : {total_enroll['Enroll_2324']:>10,}")
print(f"    2024-25 : {total_enroll['Enroll_2425']:>10,}")
print(f"    2025-26 : {total_enroll['Enroll_2526']:>10,}")
print(f"    4-yr Growth: {overall_growth:>+7.1f}%")

print(f"\n  SCHOOL HEALTH")
print(f"    Growing schools  : {growing_count:>6,}  ({growing_count/total_schools*100:.1f}%)")
print(f"    Declining schools: {declining_count:>6,}  ({declining_count/total_schools*100:.1f}%)")
print(f"    HIGH RISK        : {high_risk_count:>6,}")
print(f"    MEDIUM RISK      : {medium_risk_count:>6,}")
print(f"    Model Schools    : {model_count:>6,}")
print(f"    Closure Risk     : {closure_count:>6,}")

if avg_girls_ratio:
    print(f"\n  GENDER")
    print(f"    Avg Girls Ratio : {avg_girls_ratio:.1f}%")

print(f"\n  ALERTS")
print(f"    Critical alert types : {critical_alert_count}")
if not l5_alerts.empty:
    for _, row in l5_alerts.sort_values("Schools_Affected" if "Schools_Affected" in l5_alerts.columns else l5_alerts.columns[0], ascending=False).head(5).iterrows():
        alert_id   = row.get("Alert_ID", "")
        alert_type = row.get("Alert_Type", "")
        severity   = row.get("Severity", "")
        schools    = row.get("Schools_Affected", "?")
        print(f"      {alert_id:<4} {alert_type:<35} {severity:<10} {schools} schools")

print(f"\n  RED ZONE BLOCKS: {len(red_blocks)}")
if not red_blocks.empty:
    for _, rb in red_blocks.head(5).iterrows():
        d = rb.get("District", "")
        b = rb.get("Block", "")
        n = rb.get("Total_Schools", "")
        g = rb.get("Avg_Growth_pct", "")
        print(f"      {d} / {b}  ({n} schools, avg growth {g}%)")

if not b1_stype.empty:
    print(f"\n  SCHOOL TYPE BREAKDOWN")
    for _, row in b1_stype.iterrows():
        st  = row.get("School_Type", "")
        n   = row.get("Total_Schools", "")
        g   = row.get("Overall_Growth_pct", "")
        tag = row.get("Type_Health_Tag", "")
        print(f"      {st:<30} {n:>5} schools  {float(g):>+6.1f}%  [{tag}]")

print(f"\n  OUTPUTS WRITTEN")
print(f"    L1_STATE/  state_kpis.csv")
print(f"    L1_STATE/  state_enrollment_trend.csv")
print(f"    L5_MASTER/ master_school_profiles.csv")
print(f"    L5_MASTER/ top20_growth_schools.csv")
print(f"    L5_MASTER/ top20_decline_schools.csv")
print(f"    L5_MASTER/ top20_high_risk_schools.csv")
print(f"    L5_MASTER/ red_zone_blocks.csv")
print(f"    L5_MASTER/ same_condition_clusters.csv")
print(f"    L5_MASTER/ district_summary.csv")
print(f"    L5_MASTER/ MASTER_SUMMARY.xlsx")
print(f"    outputs/   dashboard_data.json")
print("\n" + "="*60)
print("BATCH 5 COMPLETE")
print("="*60 + "\n")
